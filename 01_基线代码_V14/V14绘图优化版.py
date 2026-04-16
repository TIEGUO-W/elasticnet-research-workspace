#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
弹性网络反演辐射源活度（鲁棒优化版 v5.1-Clean）
- 移除：冗余重启机制 (n_restarts)，依赖 GridSearchCV 原生搜索能力
- 增强：处理极端病态矩阵（条件数 1e6 ~ 1e12）
- 修复：Pylance 类型错误、None 运算问题
- 更新：导出文件增加时间戳命名 (YYYYMMDD_HHMMSS)
- 更新：每次运行使用不同随机种子，确保结果随机
- 优化：绘图字体自适应、跨平台中文兼容、400 DPI 高清输出
"""

from __future__ import annotations
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any, Union
import warnings
import sys
import platform
import math
from datetime import datetime
import numpy as np

# 设置 matplotlib 后端
import matplotlib
matplotlib.use('Agg')  # 安全后端

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib import font_manager
from sklearn.linear_model import ElasticNet
from sklearn.model_selection import GridSearchCV
from sklearn.metrics import mean_squared_error
from sklearn.exceptions import ConvergenceWarning
from tqdm import tqdm

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

warnings.filterwarnings('ignore', category=FutureWarning)
warnings.filterwarnings('ignore', category=UserWarning)
warnings.filterwarnings('ignore', category=ConvergenceWarning)

def setup_chinese_font() -> str:
    """跨平台中文字体配置 + 自动验证回退"""
    system = platform.system()
    font_map = {
        'Windows': ['Microsoft YaHei', 'SimHei', 'SimSun', 'KaiTi'],
        'Darwin':  ['PingFang SC', 'Heiti SC', 'STHeiti', 'Songti SC'],
        'Linux':   ['WenQuanYi Micro Hei', 'WenQuanYi Zen Hei', 'Noto Sans CJK SC', 'Source Han Sans CN']
    }
    candidates = font_map.get(system, font_map['Linux'])
    
    for font_name in candidates:
        try:
            font_path = font_manager.findfont(font_name)
            if font_path and 'unknown' not in font_path.lower():
                plt.rcParams['font.sans-serif'] = [font_name]
                plt.rcParams['font.family'] = 'sans-serif'
                plt.rcParams['axes.unicode_minus'] = False
                fig, ax = plt.subplots()
                ax.text(0.5, 0.5, '中', fontsize=10)
                plt.close(fig)
                print(f"✅ 中文字体已启用：{font_name}")
                return font_name
        except Exception:
            continue
            
    print("⚠️  未找到系统可用中文字体，降级为 DejaVu Sans")
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'sans-serif']
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['axes.unicode_minus'] = True
    return 'DejaVu Sans'

CHINESE_FONT = setup_chinese_font()

@dataclass
class Config:
    m: int = 8
    n: int = 10
    room_size: float = 10.0
    min_distance: float = 0.3
    sparsity_ratio: float = 0.8
    max_value: float = 3.0
    max_iter: int = 10000
    cv_folds: int = 5
    alpha_range: Tuple[float, float] = (-6, -1)
    alpha_num: int = 50
    l1_ratio_range: Tuple[float, float] = (1e-10, 1-1e-10)
    l1_ratio_num: int = 80
    run_count: int = 5
    random_seed: Optional[int] = None
    export_results: bool = True
    plot_results: bool = True
    n_bootstrap: int = 100
    output_dir: Path = Path("results")
    timestamp_str: str = field(default_factory=lambda: datetime.now().strftime("%Y%m%d_%H%M%S"))
    intercept_target_min: float = 1.0
    intercept_target_max: float = 10.0
    intercept_optimization: bool = True
    
    def __post_init__(self):
        self.output_dir.mkdir(exist_ok=True)
        if self.random_seed is None:
            self.random_seed = int(datetime.now().timestamp() * 100000) % (2**31)

def generate_sparse_solution(n: int, sparsity: float, max_val: float, 
                           rng: np.random.Generator) -> np.ndarray:
    if not 0.0 <= sparsity <= 1.0:
        raise ValueError("sparsity 必须在 [0, 1] 范围内")
    n_nz = n - int(sparsity * n)
    x = np.zeros(n)
    if n_nz > 0:
        idx = rng.choice(n, size=n_nz, replace=False)
        x[idx] = rng.uniform(0, max_val, size=n_nz)
    return x

def create_physical_system(cfg: Config, rng: np.random.Generator) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    det_pos = rng.uniform(0, cfg.room_size, size=(cfg.m, 3))
    src_pos = rng.uniform(0, cfg.room_size, size=(cfg.n, 3))
    
    A = np.zeros((cfg.m, cfg.n))
    for i in range(cfg.m):
        for j in range(cfg.n):
            dist = np.linalg.norm(det_pos[i] - src_pos[j])
            dist = max(dist, cfg.min_distance)
            A[i, j] = 1.0 / (dist ** 2)
    
    x_true = generate_sparse_solution(cfg.n, cfg.sparsity_ratio, cfg.max_value, rng)
    b = A @ x_true
    return A, x_true, b

def normalize_columns(A: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    norms = np.linalg.norm(A, axis=0)
    norms = np.where(norms == 0, 1.0, norms)
    return A / norms, norms

def safe_extract_intercept(estimator, fit_intercept: bool = True) -> float:
    if not fit_intercept:
        return 0.0
    try:
        intercept = estimator.intercept_
        if hasattr(intercept, '__len__') and len(intercept) > 0:
            return float(intercept[0])
        else:
            return float(intercept)
    except Exception:
        return 0.0

def analyze_intercept(intercept: float, b_mean: float) -> Dict[str, Union[float, str, int]]:
    intercept, b_mean = float(intercept), float(b_mean)
    ratio = abs(intercept) / abs(b_mean) if b_mean != 0 else math.inf
    ratio_pct = ratio * 100.0
    
    if ratio_pct < 1.0:
        level, color, score = "✅ 可忽略", "green", 1
    elif 1.0 <= ratio_pct < 10.0:
        level, color, score = "⭐ 轻微（最优）", "blue", 3
    else:
        level, color, score = "🔴 显著", "red", 0
        
    return {
        'value': intercept, 'b_mean': b_mean, 'ratio_pct': ratio_pct,
        'level': level, 'color': color, 'score': score,
        'interpretation': f"{level} (评分：{score}/3)"
    }

def improved_scorer(estimator, X, y):
    y_pred = estimator.predict(X)
    mse = mean_squared_error(y, y_pred)
    intercept = safe_extract_intercept(estimator, True)
    b_mean = np.mean(y)
    ratio = abs(intercept) / abs(b_mean) if b_mean != 0 else 0.0
    
    if ratio > 0.15: penalty = 0.3
    elif ratio > 0.10: penalty = 0.6
    elif ratio < 0.01: penalty = 0.9
    else: penalty = 1.0
    
    return -np.log(mse + 1e-10) * penalty

def get_alpha_range_by_condition(cond_num: float) -> Tuple[float, float]:
    if cond_num > 1e12: return (-2, 3)
    elif cond_num > 1e10: return (-3, 2)
    elif cond_num > 1e8: return (-4, 1)
    elif cond_num > 1e6: return (-5, 0)
    else: return (-6, -1)

def fit_elastic_net_robust(A: np.ndarray, b: np.ndarray, cfg: Config, 
                           fit_intercept: bool = True) -> Dict[str, Any]:
    cond_num: float = np.inf
    try: cond_num = float(np.linalg.cond(A))
    except Exception: pass
    
    alpha_range = get_alpha_range_by_condition(cond_num)
    
    # 主流程：直接执行一次高质量网格搜索
    try:
        param_grid = {
            'alpha': np.logspace(alpha_range[0], alpha_range[1], cfg.alpha_num),
            'l1_ratio': np.linspace(*cfg.l1_ratio_range, cfg.l1_ratio_num)
        }
        model = ElasticNet(
            max_iter=cfg.max_iter, fit_intercept=fit_intercept,
            positive=True, random_state=cfg.random_seed, tol=1e-6
        )
        grid = GridSearchCV(
            model, param_grid, cv=cfg.cv_folds,
            scoring=improved_scorer, n_jobs=-1, refit=True
        )
        grid.fit(A, b)
        est = grid.best_estimator_
        coef = getattr(est, 'coef_', np.zeros(A.shape[1]))
        intercept = safe_extract_intercept(est, fit_intercept)
        mse = mean_squared_error(b, est.predict(A))
        ia = analyze_intercept(intercept, b.mean())
        
        return {
            'coef': np.asarray(coef, dtype=np.float64), 'intercept': intercept,
            'params': grid.best_params_, 'model': est, 'mse': float(mse),
            'intercept_ratio': float(ia['ratio_pct']), 'condition_number': cond_num
        }
    except Exception:
        pass

    # 兜底流程：使用默认参数范围
    try:
        param_grid = {
            'alpha': np.logspace(*cfg.alpha_range, cfg.alpha_num),
            'l1_ratio': np.linspace(*cfg.l1_ratio_range, cfg.l1_ratio_num)
        }
        model = ElasticNet(
            max_iter=cfg.max_iter, fit_intercept=fit_intercept,
            positive=True, random_state=cfg.random_seed
        )
        grid = GridSearchCV(
            model, param_grid, cv=cfg.cv_folds,
            scoring='neg_mean_squared_error', n_jobs=-1
        )
        grid.fit(A, b)
        est = grid.best_estimator_
        coef = getattr(est, 'coef_', np.zeros(A.shape[1]))
        intercept = safe_extract_intercept(est, fit_intercept)
        mse = mean_squared_error(b, est.predict(A))
        ia = analyze_intercept(intercept, b.mean())
        
        return {
            'coef': np.asarray(coef, dtype=np.float64), 'intercept': intercept,
            'params': grid.best_params_, 'model': est, 'mse': float(mse),
            'intercept_ratio': float(ia['ratio_pct']), 'condition_number': cond_num
        }
    except Exception:
        return {
            'coef': np.zeros(A.shape[1], dtype=np.float64), 'intercept': 0.0,
            'params': {'alpha': 1.0, 'l1_ratio': 0.5}, 'model': None, 'mse': 1.0,
            'intercept_ratio': 0.0, 'condition_number': cond_num
        }

def run_experiment(cfg: Config) -> List[Dict[str, Any]]:
    results = []
    rng = np.random.default_rng(cfg.random_seed)
    for exp_id in tqdm(range(cfg.run_count), desc="实验进度", unit="次"):
        A, x_true, b = create_physical_system(cfg, rng)
        A_norm, col_norms = normalize_columns(A)
        res = fit_elastic_net_robust(A_norm, b, cfg, fit_intercept=True)
        x_est_norm = res['coef']
        x_est = x_est_norm / col_norms
        ia = analyze_intercept(res['intercept'], b.mean())
        
        results.append({
            'experiment_id': exp_id + 1, 'true_x': x_true, 'estimated_x': x_est,
            'intercept': res['intercept'], 'intercept_analysis': ia,
            'params': res['params'], 'mse': res['mse'],
            'condition_number': res['condition_number']
        })
        print(f"\n🔮 实验 {exp_id+1}")
        print(f"  📊 条件数：{res['condition_number']:.2e}")
        print(f"  📊 β₀={ia['value']:+.2e} |β₀/b̄|={ia['ratio_pct']:.3f}% {ia['level']}")
        print(f"  🏆 截距评分：{ia['score']}/3")
        print(f"  🔧 α={res['params']['alpha']:.1e}, l1={res['params']['l1_ratio']:.3f}")
        print(f"  📈 MSE={results[-1]['mse']:.2e}")
    return results

def _format_excel_sheet(writer, sheet_name: str, freeze_rows: int = 1, freeze_cols: int = 0):
    worksheet = writer.sheets[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except Exception: pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in worksheet[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = header_align; cell.border = thin_border
    
    if freeze_rows > 0 or freeze_cols > 0:
        freeze_cell = chr(65 + freeze_cols) + str(freeze_rows + 1)
        worksheet.freeze_panes = freeze_cell
    
    key_columns = {'MSE': 'FFC000', '|β₀/b̄| (%)': '92D050', '截距评估': '92D050', '相对误差 (%)': 'FF6B6B'}
    headers = [cell.value for cell in worksheet[1]]
    for col_idx, header in enumerate(headers, start=1):
        if header in key_columns:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).fill = PatternFill(start_color=key_columns[header], end_color=key_columns[header], fill_type='solid')
                worksheet.cell(row=row, column=col_idx).alignment = Alignment(horizontal='center')
    
    numeric_cols = ['MSE', '截距 β₀', '|β₀/b̄| (%)', 'alpha', 'l1_ratio', '真实活度', '估计活度', '绝对误差', '相对误差 (%)']
    for col_idx, header in enumerate(headers, start=1):
        if header in numeric_cols:
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=col_idx).alignment = Alignment(horizontal='right')

def export_results(results: List[Dict[str, Any]], cfg: Config) -> Optional[Path]:
    if not HAS_OPENPYXL:
        print("⚠️  未安装 openpyxl，跳过 Excel 导出"); return None
    output_dir = cfg.output_dir; output_dir.mkdir(exist_ok=True)
    records_wide = []
    for i, r in enumerate(results):
        record = {
            '实验序号': i + 1, 'MSE': f'{r["mse"]:.2e}', '截距 β₀': f'{r["intercept"]:+.2e}',
            '|β₀/b̄| (%)': f'{r["intercept_analysis"]["ratio_pct"]:.3f}', '截距评估': r['intercept_analysis']['level'],
            '条件数': f'{r["condition_number"]:.2e}', 'alpha': f'{r["params"]["alpha"]:.1e}', 'l1_ratio': f'{r["params"]["l1_ratio"]:.4f}'
        }
        for j, val in enumerate(r['true_x']): record[f'真实_x[{j}]'] = f'{val:.4f}'
        for j, val in enumerate(r['estimated_x']): record[f'估计_x[{j}]'] = f'{val:.4f}'
        records_wide.append(record)
    df_wide = pd.DataFrame(records_wide)
    csv_wide = output_dir / f"results_wide_{cfg.timestamp_str}.csv"
    df_wide.to_csv(csv_wide, index=False, encoding='utf-8-sig')
    print(f"💾 CSV 已导出：{csv_wide}")
    
    excel_path = output_dir / f"results_{cfg.timestamp_str}.xlsx"
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_wide.to_excel(writer, sheet_name='概览_宽表', index=False)
            _format_excel_sheet(writer, '概览_宽表', freeze_rows=1, freeze_cols=1)
        print(f"📊 Excel 已导出：{excel_path}"); return excel_path
    except Exception as e:
        print(f"⚠️  Excel 导出失败：{e}"); return None

def plot_results_subplots(results: List[Dict[str, Any]], cfg: Config):
    """自适应字体 + 跨平台兼容 + 400 DPI 出版级绘图"""
    if not results: return
    
    n_feat = len(results[0]['true_x'])
    n_exp = len(results)
    c = np.arange(1, n_feat + 1)
    
    fig_height = max(6.5, n_exp * 5.2)
    fig, axes = plt.subplots(n_exp, 1, figsize=(15, fig_height), dpi=100)
    if n_exp == 1: axes = [axes]
    
    scale = min(1.0, 6 / max(n_exp, 1))
    FS = {'title': 13*scale, 'label': 11*scale, 'legend': 9*scale, 'tick': 9*scale}
    
    colors = ['#2E86AB', '#E94F37', '#44AF69', '#F3722C', '#A8DADC', '#011627']
    markers = ['o', 's', '^', 'd', 'v', 'p']
    font_prop = font_manager.FontProperties(family=CHINESE_FONT)
    
    for idx, (ax, res) in enumerate(zip(axes, results)):
        true_x, est_x = res['true_x'], res['estimated_x']
        mse, ia, cond = res['mse'], res['intercept_analysis'], res['condition_number']
        color, marker = colors[idx % len(colors)], markers[idx % len(markers)]
        
        ax.plot(c, true_x, marker='o', linestyle='-', color='#000000', 
               label='真实解', linewidth=2.0, markersize=7, markerfacecolor='white', markeredgecolor='black', markeredgewidth=1.5)
        ax.plot(c, est_x, marker=marker, linestyle='--', color=color,
               label='估计解', linewidth=1.8, markersize=6, alpha=0.9)
        
        title_txt = f"实验 {idx+1} | MSE = {mse:.2e}\n截距 β₀ = {ia['value']:+.2e} (|β₀/ȳ| = {ia['ratio_pct']:.2f}%) - {ia['level']}\n条件数 κ = {cond:.2e}"
        ax.set_title(title_txt, fontsize=FS['title'], fontweight='bold', pad=10, fontproperties=font_prop)
        ax.set_xlabel('源索引 (Source Index)', fontsize=FS['label'], fontproperties=font_prop)
        ax.set_ylabel('活度值 (Activity)', fontsize=FS['label'], fontproperties=font_prop)
        ax.tick_params(axis='both', labelsize=FS['tick'])
        ax.set_xticks(c); ax.set_xticklabels([str(int(i)) for i in c])
        
        ax.grid(True, linestyle=':', linewidth=0.7, alpha=0.6, color='#888888')
        ax.axhline(y=0, color='#CCCCCC', linestyle='-', linewidth=0.4, alpha=0.5)
        ax.legend(fontsize=FS['legend'], frameon=True, fancybox=True, framealpha=0.9, edgecolor='#AAAAAA', prop=font_prop, loc='upper right')
        
        for spine in ax.spines.values(): spine.set_edgecolor('#333333'); spine.set_linewidth(0.8)
    
    fig.suptitle(f'弹性网络反演结果对比（{n_exp} 次独立实验）· 鲁棒优化版 v5.1', 
                fontsize=15, fontweight='bold', y=0.995, fontproperties=font_prop)
    plt.tight_layout(rect=[0, 0.02, 1, 0.98])
    
    plot_path = cfg.output_dir / f"results_subplots_{cfg.timestamp_str}.png"
    plt.savefig(plot_path, dpi=400, bbox_inches='tight', pad_inches=0.1, facecolor='white')
    print(f"🖼️  高清子图已保存：{plot_path} (400 DPI)")
    plt.close(fig)

def print_summary(results: List[Dict[str, Any]]):
    if not results: return
    print("\n" + "═"*70)
    print("📈 实验总结（鲁棒优化版）")
    print("═"*70)
    mses = [r['mse'] for r in results]
    ratios = [r['intercept_analysis']['ratio_pct'] for r in results]
    cond_nums = [r['condition_number'] for r in results]
    print(f"🔹 实验次数：{len(results)}")
    print(f"🔹 条件数范围：[{min(cond_nums):.2e}, {max(cond_nums):.2e}]")
    print(f"🔹 MSE: 均值={np.mean(mses):.2e}, 标准差={np.std(mses):.2e}")
    print(f"🔹 |β₀/b̄| 比例：均值={np.mean(ratios):.3f}%, 范围=[{min(ratios):.3f}%, {max(ratios):.3f}%]")
    n_optimal = sum(1 for r in ratios if 1.0 <= r < 10.0)
    print(f"\n🔹 截距最优比例（1%~10%）：{n_optimal}/{len(results)} ({n_optimal/len(results)*100:.1f}%)")
    mse_cv = np.std(mses) / np.mean(mses) if np.mean(mses) > 0 else 0.0
    stability = "✅ 优秀" if mse_cv < 0.5 else "⚠️ 良好" if mse_cv < 1.0 else "❌ 较差"
    print(f"🔹 稳定性：MSE 变异系数={mse_cv:.2f} {stability}")
    print("═"*70 + "\n")

def main():
    print("="*70)
    print("🚀 弹性网络反演辐射源活度实验（鲁棒优化版 v5.1-Clean）")
    print("="*70)
    cfg = Config(
        run_count=5, export_results=True, plot_results=True,
        n_bootstrap=100, cv_folds=5, intercept_target_min=1.0, intercept_target_max=10.0,
        intercept_optimization=True
    )
    print(f"📋 配置：m={cfg.m}, n={cfg.n}, runs={cfg.run_count}")
    print(f"🎲 随机种子：{cfg.random_seed} (每次运行不同)")
    print(f"🕒 时间戳：{cfg.timestamp_str}")
    print(f"🎯 截距优化目标：{cfg.intercept_target_min}% ~ {cfg.intercept_target_max}%")
    print(f"🔍 条件数自适应：启用")
    print(f"📦 功能：导出={cfg.export_results}, 绘图={cfg.plot_results}")
    print(f"✅ 中文字体：{CHINESE_FONT}")
    print(f"✅ Excel 导出：{'可用' if HAS_OPENPYXL else '不可用'}")
    print("="*70 + "\n")
    try:
        results = run_experiment(cfg)
    except Exception as e:
        print(f"\n❌ 实验失败：{e}")
        import traceback; traceback.print_exc()
        sys.exit(1)
    print_summary(results)
    if cfg.export_results:
        try: export_results(results, cfg)
        except Exception as e: print(f"⚠️  导出失败：{e}")
    if cfg.plot_results:
        try: plot_results_subplots(results, cfg)
        except Exception as e: print(f"⚠️  绘图失败：{e}")
    print("✨ 实验完成！")

if __name__ == "__main__":
    main()